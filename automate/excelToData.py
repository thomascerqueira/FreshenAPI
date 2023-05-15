import argparse
import openpyxl
import requests
import os

local_url = "http://localhost:4500"
gestion_article = "/V2/gestion_article/"

class NoPhoto(Exception):
    pass

class Article:
    brand: str
    article: str
    coton: float
    prix: float
    eau: float
    
    def add(self, url: str):
        x = requests.post(
            url + gestion_article + self.brand + "/" + self.article,
            json={
                "cost": self.prix,
                "coton": self.coton,
                "water": self.eau
            }
        )
        
        if (x.status_code != 200):
            print("{}:\n\t{} {}".format(self.brand + "->" + self.article, x.status_code, x.text))
        else:
            print("Article added: {}->{}".format(self.brand, self.article))
    
    def __init__(self, brand, article, coton, prix, eau) -> None:
        self.brand = brand
        self.article = article
        self.coton = coton
        self.prix = prix
        self.eau = eau
    
    def __str__(self):
        return "Brand: {}\nArticle: {}\nCoton: {}\nPrix: {}\nEau: {}\n".format(brand, article, coton, prix, eau)

def addBrand(name: str, url: str):
    file_path = ""
        
    for file in os.listdir("photos"):
        if file.startswith(name):
            file_path = "photos/" + file
            break
    if file_path == "":
        raise NoPhoto("No photo")
    files = {'upload_file': open(file_path,'rb')}
    x = requests.post(
        url + gestion_article + "brand",
        json= {
            "brand": name
        },
        files=files
    )

    if (x.status_code != 200):
        print("{}:\n\t{} {}".format(name, x.status_code, x.text))
    else:
        print("Brand added: {}".format(name))

def updateBrand(url: str, data):
    value = []
    
    for val in data:
        if val not in value and val != None:
            value.append(val)
            
    for val in value:
        try:
            addBrand(val, url)
        except NoPhoto as e:
          print("No photo for: {}, ignoring".format(val))
    return

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Take all data from the excel and add to server")
    
    parser.add_argument('filename')
    parser.add_argument('-b', '--brand', action='store_true')
    parser.add_argument('-a', '--article', action='store_true')
    parser.add_argument('-l', '--local', action='store_false')
    
    args = parser.parse_args()
    
    url = local_url if args.local else ""
    
    ws = openpyxl.load_workbook(args.filename).active
    
    
    if (args.brand):
        brands = [ws.cell(row=i,column=2).value for i in range(2,ws.max_row+1)]
        updateBrand(url, brands)
    
    if args.article:
        for row in range(2, ws.max_row + 1):
            article = ws.cell(row, 1).value
            brand = ws.cell(row, 2).value
            prix = ws.cell(row, 3).value
            coton = ws.cell(row, 5).value
            eau = ws.cell(row, 8).value
            
            if brand == None or article == None:
                continue
            try:
                prix = float(prix)
            except:
                prix = 0
        
            try:
                coton = float(coton)
            except:
                coton = 0
                
            try:
                eau = float(eau)
            except:
                eau = 0
            article = Article(brand, article, coton, prix, eau)
            
            article.add(url)